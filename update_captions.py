#!/usr/bin/env python3
"""
update_captions.py
Reads captions.xlsx and writes captions.json for the Vibhav Lal portfolio website.

Usage:
    python3 update_captions.py

The Excel file must have these columns (row 1 = headers):
    category | filename | title | description

After running, open your browser to see the updated hover text.
If the site is on GitHub Pages, commit and push captions.json.
"""

import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip3 install openpyxl")
    sys.exit(1)

BASE_DIR = Path(__file__).parent
XLSX_FILE = BASE_DIR / "captions.xlsx"
JSON_FILE = BASE_DIR / "captions.json"

VALID_CATEGORIES = {"lego", "sketches", "rubiks", "woodwork"}


def xlsx_to_json():
    if not XLSX_FILE.exists():
        print(f"ERROR: {XLSX_FILE} not found.")
        sys.exit(1)

    wb = openpyxl.load_workbook(XLSX_FILE)
    ws = wb.active

    headers = [str(cell.value).strip().lower() if cell.value else "" for cell in ws[1]]
    required = {"category", "filename", "title", "description"}
    missing = required - set(headers)
    if missing:
        print(f"ERROR: Missing columns in Excel: {missing}")
        print(f"Found columns: {headers}")
        sys.exit(1)

    col = {name: headers.index(name) for name in required}
    captions = {cat: {} for cat in VALID_CATEGORIES}
    skipped = 0

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        category = str(row[col["category"]] or "").strip().lower()
        filename = str(row[col["filename"]] or "").strip()
        title = str(row[col["title"]] or "").strip()
        description = str(row[col["description"]] or "").strip()

        if not category and not filename:
            continue  # blank row

        if category not in VALID_CATEGORIES:
            print(f"  Row {row_num}: unknown category '{category}' — skipped")
            skipped += 1
            continue

        filename = filename.replace(".jpg", "").replace(".JPG", "").replace(".jpeg", "")
        captions[category][filename] = {"title": title, "description": description}

    with open(JSON_FILE, "w") as f:
        json.dump(captions, f, indent=2)

    total = sum(len(v) for v in captions.values())
    print(f"Done. {total} captions written to captions.json ({skipped} rows skipped).")
    if skipped == 0:
        print("Next step: commit and push captions.json to GitHub to update the live site.")


if __name__ == "__main__":
    xlsx_to_json()
